from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import json, os

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'zawadibora-farm-secret-key-2026')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///zawadibora.db').replace('postgres://', 'postgresql://')
db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# ===== MODELS =====
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='viewer')
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class RawMaterial(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    unit = db.Column(db.String(20), default='kg')
    bags = db.Column(db.Text, default='[]')
    current_stock = db.Column(db.Float, default=0)
    unit_cost = db.Column(db.Float, default=0)
    reorder_level = db.Column(db.Float, default=0)
    description = db.Column(db.Text, default='')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    unit = db.Column(db.String(20), default='bag')
    bags = db.Column(db.Text, default='[]')
    current_stock = db.Column(db.Float, default=0)
    price = db.Column(db.Float, default=0)
    description = db.Column(db.Text, default='')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Supplier(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    contact = db.Column(db.String(200), default='')
    phone = db.Column(db.String(50), default='')
    email = db.Column(db.String(100), default='')
    address = db.Column(db.Text, default='')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Purchase(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    raw_material_id = db.Column(db.Integer, db.ForeignKey('raw_material.id'))
    supplier_id = db.Column(db.Integer, db.ForeignKey('supplier.id'), nullable=True)
    supplier_name = db.Column(db.String(200), default='')
    quantity = db.Column(db.Float, default=0)
    unit_price = db.Column(db.Float, default=0)
    total_cost = db.Column(db.Float, default=0)
    paid_amount = db.Column(db.Float, default=0)
    balance = db.Column(db.Float, default=0)
    payment_method = db.Column(db.String(50), default='Cash')
    payment_date = db.Column(db.String(20), default='')
    invoice_no = db.Column(db.String(100), default='')
    date = db.Column(db.String(20), default='')
    notes = db.Column(db.Text, default='')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Production(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'))
    quantity_produced = db.Column(db.Float, default=0)
    ingredients = db.Column(db.Text, default='[]')
    date = db.Column(db.String(20), default='')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Sale(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'))
    quantity = db.Column(db.Float, default=0)
    unit_price = db.Column(db.Float, default=0)
    total_amount = db.Column(db.Float, default=0)
    paid_amount = db.Column(db.Float, default=0)
    balance = db.Column(db.Float, default=0)
    customer_name = db.Column(db.String(200), default='Walk-in')
    date = db.Column(db.String(20), default='')
    payment_method = db.Column(db.String(50), default='Cash')
    payment_date = db.Column(db.String(20), default='')
    notes = db.Column(db.Text, default='')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# ===== HELPERS =====
BAG_SIZES = [70, 50, 35, 20, 10]

def fmt_num(n):
    try: return f"{n:,.0f}" if n == int(n) else f"{n:,.2f}"
    except: return "0"

def fmt_money(n):
    try: return f"KSh {n:,.2f}"
    except: return "KSh 0.00"

def today_str():
    return datetime.now().strftime('%Y-%m-%d')

def parse_bags(bags_json):
    if not bags_json: return []
    try: return json.loads(bags_json) if isinstance(bags_json, str) else bags_json
    except: return []

def bags_total(bags):
    return sum(b.get('qty', 0) for b in bags)

def bags_to_kg(bags):
    return sum(b.get('size', 0) * b.get('qty', 0) for b in bags)

def bag_breakdown(bags):
    if not bags: return ''
    return ' + '.join(f"{b.get('qty',0)}x{b.get('size',0)}kg" for b in bags)

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if current_user.role != 'admin':
            flash('Only the admin can perform this action.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

@app.context_processor
def inject_globals():
    return dict(today_str=today_str, fmt_num=fmt_num, fmt_money=fmt_money,
        parse_bags=parse_bags, bags_total=bags_total, bags_to_kg=bags_to_kg,
        bag_breakdown=bag_breakdown, BAG_SIZES=BAG_SIZES)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ===== AUTH =====
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form['username']).first()
        if user and user.check_password(request.form['password']):
            login_user(user)
            return redirect(url_for('dashboard'))
        flash('Invalid username or password', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ===== DASHBOARD =====
@app.route('/')
@login_required
def dashboard():
    rm = RawMaterial.query.all()
    prod = Product.query.all()
    purchases = Purchase.query.all()
    sales = Sale.query.all()
    total_rm = sum(r.current_stock * (r.unit_cost or 0) for r in rm)
    total_prod = sum(p.current_stock * (p.price or 0) for p in prod)
    total_purchases = sum(p.total_cost for p in purchases)
    total_sales = sum(s.total_amount for s in sales)
    total_from_customers = sum(s.paid_amount or 0 for s in sales)
    owed_by_customers = sum(s.balance or s.total_amount for s in sales)
    paid_to_suppliers = sum(p.paid_amount or 0 for p in purchases)
    owing_to_suppliers = sum(p.balance or p.total_cost for p in purchases)
    low_stock = [r for r in rm if r.current_stock <= r.reorder_level]
    return render_template('dashboard.html', rm=rm, prod=prod,
        total_rm=total_rm, total_prod=total_prod,
        total_purchases=total_purchases, total_sales=total_sales,
        total_from_customers=total_from_customers,
        owed_by_customers=owed_by_customers,
        paid_to_suppliers=paid_to_suppliers,
        owing_to_suppliers=owing_to_suppliers,
        low_stock=low_stock, fmt_num=fmt_num, fmt_money=fmt_money,
        parse_bags=parse_bags, bags_total=bags_total, bags_to_kg=bags_to_kg,
        bag_breakdown=bag_breakdown)

# ===== RAW MATERIALS =====
@app.route('/raw-materials')
@login_required
def raw_materials():
    items = RawMaterial.query.all()
    return render_template('raw_materials.html', items=items, fmt_num=fmt_num,
        fmt_money=fmt_money, parse_bags=parse_bags, bags_total=bags_total,
        bags_to_kg=bags_to_kg, bag_breakdown=bag_breakdown, BAG_SIZES=BAG_SIZES)

@app.route('/raw-materials/save', methods=['POST'])
@login_required
@admin_required
def raw_material_save():
    data = request.form
    if data['unit'] == 'bag':
        bags = json.loads(data.get('bags_json', '[]'))
        if not bags: flash('Add at least one bag size', 'danger'); return redirect(url_for('raw_materials'))
        stock = bags_total(bags)
    else:
        bags = []
        stock = float(data.get('current_stock', 0))
    if data.get('id'):
        item = RawMaterial.query.get(int(data['id']))
        if item:
            item.name = data['name']; item.unit = data['unit']
            item.bags = json.dumps(bags); item.current_stock = stock
            item.unit_cost = float(data.get('unit_cost', 0))
            item.reorder_level = float(data.get('reorder_level', 0))
            item.description = data.get('description', '')
    else:
        item = RawMaterial(name=data['name'], unit=data['unit'],
            bags=json.dumps(bags), current_stock=stock,
            unit_cost=float(data.get('unit_cost', 0)),
            reorder_level=float(data.get('reorder_level', 0)),
            description=data.get('description', ''))
        db.session.add(item)
    db.session.commit()
    return redirect(url_for('raw_materials'))

@app.route('/raw-materials/delete/<int:id>')
@login_required
@admin_required
def raw_material_delete(id):
    item = RawMaterial.query.get(id)
    if item: db.session.delete(item); db.session.commit()
    return redirect(url_for('raw_materials'))

@app.route('/raw-materials/edit/<int:id>')
@login_required
def raw_material_edit(id):
    item = RawMaterial.query.get(id)
    if not item: return jsonify({})
    return jsonify({'id': item.id, 'name': item.name, 'unit': item.unit,
        'bags': json.loads(item.bags or '[]'), 'current_stock': item.current_stock,
        'unit_cost': item.unit_cost, 'reorder_level': item.reorder_level,
        'description': item.description or ''})

# ===== PRODUCTS =====
@app.route('/products')
@login_required
def products():
    items = Product.query.all()
    return render_template('products.html', items=items, fmt_num=fmt_num,
        fmt_money=fmt_money, parse_bags=parse_bags, bags_total=bags_total,
        bags_to_kg=bags_to_kg, bag_breakdown=bag_breakdown, BAG_SIZES=BAG_SIZES)

@app.route('/products/save', methods=['POST'])
@login_required
@admin_required
def product_save():
    data = request.form
    if data['unit'] == 'bag':
        bags = json.loads(data.get('bags_json', '[]'))
        if not bags: flash('Add at least one bag size', 'danger'); return redirect(url_for('products'))
        stock = bags_total(bags)
    else:
        bags = []
        stock = float(data.get('current_stock', 0))
    if data.get('id'):
        item = Product.query.get(int(data['id']))
        if item:
            item.name = data['name']; item.unit = data['unit']
            item.bags = json.dumps(bags); item.current_stock = stock
            item.price = float(data.get('price', 0))
            item.description = data.get('description', '')
    else:
        item = Product(name=data['name'], unit=data['unit'],
            bags=json.dumps(bags), current_stock=stock,
            price=float(data.get('price', 0)),
            description=data.get('description', ''))
        db.session.add(item)
    db.session.commit()
    return redirect(url_for('products'))

@app.route('/products/delete/<int:id>')
@login_required
@admin_required
def product_delete(id):
    item = Product.query.get(id)
    if item: db.session.delete(item); db.session.commit()
    return redirect(url_for('products'))

@app.route('/products/edit/<int:id>')
@login_required
def product_edit(id):
    item = Product.query.get(id)
    if not item: return jsonify({})
    return jsonify({'id': item.id, 'name': item.name, 'unit': item.unit,
        'bags': json.loads(item.bags or '[]'), 'current_stock': item.current_stock,
        'price': item.price, 'description': item.description or ''})

# ===== SUPPLIERS =====
@app.route('/suppliers')
@login_required
def suppliers():
    items = Supplier.query.all()
    return render_template('suppliers.html', items=items)

@app.route('/suppliers/save', methods=['POST'])
@login_required
@admin_required
def supplier_save():
    data = request.form
    if data.get('id'):
        s = Supplier.query.get(int(data['id']))
        if s:
            s.name = data['name']; s.contact = data.get('contact', '')
            s.phone = data.get('phone', ''); s.email = data.get('email', '')
            s.address = data.get('address', '')
    else:
        s = Supplier(name=data['name'], contact=data.get('contact',''),
            phone=data.get('phone',''), email=data.get('email',''),
            address=data.get('address',''))
        db.session.add(s)
    db.session.commit()
    return redirect(url_for('suppliers'))

@app.route('/suppliers/delete/<int:id>')
@login_required
@admin_required
def supplier_delete(id):
    s = Supplier.query.get(id)
    if s: db.session.delete(s); db.session.commit()
    return redirect(url_for('suppliers'))

@app.route('/suppliers/edit/<int:id>')
@login_required
def supplier_edit(id):
    s = Supplier.query.get(id)
    if not s: return jsonify({})
    return jsonify({'id': s.id, 'name': s.name, 'contact': s.contact, 'phone': s.phone, 'email': s.email, 'address': s.address})

# ===== PURCHASES =====
@app.route('/purchases')
@login_required
def purchases():
    items = Purchase.query.all()
    rm = RawMaterial.query.all()
    suppliers = Supplier.query.all()
    return render_template('purchases.html', items=items, rm=rm,
        suppliers=suppliers, fmt_num=fmt_num, fmt_money=fmt_money)

@app.route('/purchases/save', methods=['POST'])
@login_required
@admin_required
def purchase_save():
    data = request.form
    qty = float(data.get('quantity', 0))
    price = float(data.get('unit_price', 0))
    total = qty * price
    paid = min(float(data.get('paid_amount', 0)), total)
    rm_id = int(data.get('raw_material_id'))
    p = Purchase(raw_material_id=rm_id,
        supplier_id=data.get('supplier_id') and int(data['supplier_id']) or None,
        supplier_name=data.get('supplier_name', ''),
        quantity=qty, unit_price=price, total_cost=total,
        paid_amount=paid, balance=total-paid,
        payment_method=data.get('payment_method', 'Cash'),
        payment_date=data.get('payment_date', today_str()),
        invoice_no=data.get('invoice_no', ''),
        date=data.get('date', today_str()),
        notes=data.get('notes', ''))
    db.session.add(p)
    mat = RawMaterial.query.get(rm_id)
    if mat:
        mat.current_stock = (mat.current_stock or 0) + qty
    db.session.commit()
    return redirect(url_for('purchases'))

@app.route('/purchases/data/<int:id>')
@login_required
def purchase_data(id):
    p = Purchase.query.get(id)
    if not p: return jsonify({})
    return jsonify({'id': p.id, 'supplier_name': p.supplier_name,
        'total_cost': p.total_cost, 'paid_amount': p.paid_amount, 'balance': p.balance})

@app.route('/purchases/pay', methods=['POST'])
@login_required
@admin_required
def purchase_pay():
    data = request.form
    pid = int(data.get('purchase_id'))
    amount = float(data.get('payment_amount', 0))
    p = Purchase.query.get(pid)
    if p and amount > 0:
        already = p.paid_amount or 0
        max_allowed = p.total_cost - already
        if amount > max_allowed + 0.01:
            flash(f'Overpayment! Max allowed: {fmt_money(max_allowed)}', 'danger')
            return redirect(url_for('purchases'))
        p.paid_amount = already + amount
        p.balance = p.total_cost - p.paid_amount
        p.payment_method = data.get('payment_method', p.payment_method)
        p.payment_date = data.get('payment_date', today_str())
        db.session.commit()
    return redirect(url_for('purchases'))

@app.route('/purchases/delete/<int:id>')
@login_required
@admin_required
def purchase_delete(id):
    p = Purchase.query.get(id)
    if p:
        mat = RawMaterial.query.get(p.raw_material_id)
        if mat: mat.current_stock = max(0, (mat.current_stock or 0) - p.quantity)
        db.session.delete(p)
        db.session.commit()
    return redirect(url_for('purchases'))

# ===== PRODUCTION =====
@app.route('/production')
@login_required
def production():
    items = Production.query.all()
    products = Product.query.all()
    rm = RawMaterial.query.all()
    return render_template('production.html', items=items, products=products,
        rm=rm, fmt_num=fmt_num, parse_bags=parse_bags, BAG_SIZES=BAG_SIZES)

@app.route('/production/save', methods=['POST'])
@login_required
@admin_required
def production_save():
    data = request.form
    product_id = int(data.get('product_id'))
    qty = float(data.get('quantity_produced', 0))
    ingredients = json.loads(data.get('ingredients_json', '[]'))
    # Check stock
    for ing in ingredients:
        mat = RawMaterial.query.get(ing.get('raw_material_id'))
        if mat and ing.get('qty', 0) > mat.current_stock:
            flash(f'Insufficient {mat.name}!', 'danger')
            return redirect(url_for('production'))
    # Deduct raw materials
    for ing in ingredients:
        mat = RawMaterial.query.get(ing.get('raw_material_id'))
        if mat:
            mat.current_stock = max(0, (mat.current_stock or 0) - ing['qty'])
    # Add product
    prod = Product.query.get(product_id)
    if prod:
        prod.current_stock = (prod.current_stock or 0) + qty
    p = Production(product_id=product_id, quantity_produced=qty,
        ingredients=json.dumps(ingredients), date=data.get('date', today_str()))
    db.session.add(p)
    db.session.commit()
    return redirect(url_for('production'))

@app.route('/production/delete/<int:id>')
@login_required
@admin_required
def production_delete(id):
    p = Production.query.get(id)
    if p:
        ings = json.loads(p.ingredients or '[]')
        for ing in ings:
            mat = RawMaterial.query.get(ing.get('raw_material_id'))
            if mat: mat.current_stock = (mat.current_stock or 0) + ing.get('qty', 0)
        prod = Product.query.get(p.product_id)
        if prod: prod.current_stock = max(0, (prod.current_stock or 0) - p.quantity_produced)
        db.session.delete(p)
        db.session.commit()
    return redirect(url_for('production'))

# ===== SALES =====
@app.route('/sales')
@login_required
def sales():
    items = Sale.query.all()
    products = Product.query.all()
    return render_template('sales.html', items=items, products=products,
        fmt_num=fmt_num, fmt_money=fmt_money)

@app.route('/sales/save', methods=['POST'])
@login_required
@admin_required
def sale_save():
    data = request.form
    product_id = int(data.get('product_id'))
    qty = float(data.get('quantity', 0))
    price = float(data.get('unit_price', 0))
    total = qty * price
    paid = min(float(data.get('paid_amount', 0)), total)
    prod = Product.query.get(product_id)
    if not prod or qty > prod.current_stock:
        flash('Insufficient stock!', 'danger')
        return redirect(url_for('sales'))
    s = Sale(product_id=product_id, quantity=qty, unit_price=price,
        total_amount=total, paid_amount=paid, balance=total-paid,
        customer_name=data.get('customer_name', 'Walk-in'),
        date=data.get('date', today_str()),
        payment_method=data.get('payment_method', 'Cash'),
        payment_date=data.get('payment_date', today_str()),
        notes=data.get('notes', ''))
    prod.current_stock = max(0, (prod.current_stock or 0) - qty)
    db.session.add(s)
    db.session.commit()
    return redirect(url_for('sales'))

@app.route('/sales/data/<int:id>')
@login_required
def sale_data(id):
    s = Sale.query.get(id)
    if not s: return jsonify({})
    return jsonify({'id': s.id, 'customer_name': s.customer_name,
        'total_amount': s.total_amount, 'paid_amount': s.paid_amount, 'balance': s.balance})

@app.route('/sales/pay', methods=['POST'])
@login_required
@admin_required
def sale_pay():
    data = request.form
    sid = int(data.get('sale_id'))
    amount = float(data.get('payment_amount', 0))
    s = Sale.query.get(sid)
    if s and amount > 0:
        already = s.paid_amount or 0
        max_allowed = s.total_amount - already
        if amount > max_allowed + 0.01:
            flash(f'Overpayment! Max allowed: {fmt_money(max_allowed)}', 'danger')
            return redirect(url_for('sales'))
        s.paid_amount = already + amount
        s.balance = s.total_amount - s.paid_amount
        s.payment_method = data.get('payment_method', s.payment_method)
        s.payment_date = data.get('payment_date', today_str())
        db.session.commit()
    return redirect(url_for('sales'))

@app.route('/sales/delete/<int:id>')
@login_required
@admin_required
def sale_delete(id):
    s = Sale.query.get(id)
    if s:
        prod = Product.query.get(s.product_id)
        if prod: prod.current_stock = (prod.current_stock or 0) + s.quantity
        db.session.delete(s)
        db.session.commit()
    return redirect(url_for('sales'))

# ===== PAYMENTS =====
@app.route('/payments')
@login_required
def payments():
    purchases = Purchase.query.all()
    sales = Sale.query.all()
    # Compute summaries in Python
    supplier_summary = {}
    for p in purchases:
        key = p.supplier_name or 'Unknown'
        if key not in supplier_summary:
            supplier_summary[key] = {'count': 0, 'total': 0, 'paid': 0, 'balance': 0}
        supplier_summary[key]['count'] += 1
        supplier_summary[key]['total'] += p.total_cost
        supplier_summary[key]['paid'] += p.paid_amount or 0
        supplier_summary[key]['balance'] += p.balance or p.total_cost
    customer_summary = {}
    for s in sales:
        key = s.customer_name or 'Walk-in'
        if key not in customer_summary:
            customer_summary[key] = {'count': 0, 'total': 0, 'paid': 0, 'balance': 0}
        customer_summary[key]['count'] += 1
        customer_summary[key]['total'] += s.total_amount
        customer_summary[key]['paid'] += s.paid_amount or 0
        customer_summary[key]['balance'] += s.balance or s.total_amount
    return render_template('payments.html', purchases=purchases, sales=sales,
        supplier_summary=supplier_summary, customer_summary=customer_summary)

# ===== REPORTS =====
@app.route('/reports')
@login_required
def reports():
    purchases = Purchase.query.all()
    sales = Sale.query.all()
    production = Production.query.all()
    # Monthly aggregation
    monthly_purchases = {}
    for p in purchases:
        m = (p.date or '')[:7]
        if m: monthly_purchases[m] = monthly_purchases.get(m, 0) + p.total_cost
    monthly_sales = {}
    for s in sales:
        m = (s.date or '')[:7]
        if m: monthly_sales[m] = monthly_sales.get(m, 0) + s.total_amount
    return render_template('reports.html', purchases=purchases, sales=sales,
        production=production, monthly_purchases=monthly_purchases,
        monthly_sales=monthly_sales)

@app.route('/export')
@login_required
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        flash('openpyxl not installed. Run: pip install openpyxl', 'danger')
        return redirect(url_for('reports'))
    wb = openpyxl.Workbook()
    # Raw Materials
    ws = wb.active; ws.title = 'Raw Materials'
    ws.append(['Name', 'Unit', 'Stock', 'Bag Breakdown', 'Total Kg', 'Cost/Unit', 'Value', 'Reorder'])
    for r in RawMaterial.query.all():
        bags = parse_bags(r.bags)
        bd = bag_breakdown(bags)
        kg = bags_to_kg(bags) if r.unit == 'bag' else (r.current_stock*1000 if r.unit=='tonne' else (r.current_stock if r.unit=='kg' else 0))
        ws.append([r.name, r.unit, r.current_stock, bd, kg, r.unit_cost, r.current_stock*r.unit_cost, r.reorder_level])
    # Products
    ws2 = wb.create_sheet('Products')
    ws2.append(['Name', 'Unit', 'Stock', 'Bag Breakdown', 'Total Kg', 'Price', 'Value'])
    for p in Product.query.all():
        bags = parse_bags(p.bags)
        bd = bag_breakdown(bags)
        kg = bags_to_kg(bags) if p.unit == 'bag' else (p.current_stock*1000 if p.unit=='tonne' else (p.current_stock if p.unit=='kg' else 0))
        ws2.append([p.name, p.unit, p.current_stock, bd, kg, p.price, p.current_stock*p.price])
    # Purchases
    ws3 = wb.create_sheet('Purchases')
    ws3.append(['Date', 'Material', 'Supplier', 'Qty', 'Unit Price', 'Total', 'Paid', 'Balance'])
    for p in Purchase.query.all():
        mat = RawMaterial.query.get(p.raw_material_id)
        ws3.append([p.date, mat.name if mat else 'Unknown', p.supplier_name, p.quantity, p.unit_price, p.total_cost, p.paid_amount, p.balance])
    # Sales
    ws4 = wb.create_sheet('Sales')
    ws4.append(['Date', 'Product', 'Customer', 'Qty', 'Unit Price', 'Total', 'Paid', 'Balance'])
    for s in Sale.query.all():
        prod = Product.query.get(s.product_id)
        ws4.append([s.date, prod.name if prod else 'Unknown', s.customer_name, s.quantity, s.unit_price, s.total_amount, s.paid_amount, s.balance])
    # Production
    ws5 = wb.create_sheet('Production')
    ws5.append(['Date', 'Product', 'Qty Produced', 'Ingredients'])
    for p in Production.query.all():
        prod = Product.query.get(p.product_id)
        ings = json.loads(p.ingredients or '[]')
        ing_str = '; '.join(f"{i.get('name','')}: {i.get('qty',0)}" for i in ings)
        ws5.append([p.date, prod.name if prod else 'Unknown', p.quantity_produced, ing_str])
    # Summary
    ws6 = wb.create_sheet('Summary')
    ws6.append(['Metric', 'Value'])
    ws6.append(['Total Purchases', sum(p.total_cost for p in Purchase.query.all())])
    ws6.append(['Total Sales', sum(s.total_amount for s in Sale.query.all())])
    ws6.append(['Gross Margin', sum(s.total_amount for s in Sale.query.all()) - sum(p.total_cost for p in Purchase.query.all())])

    filename = f'Zawadibora_Report_{today_str()}.xlsx'
    filepath = os.path.join(app.instance_path or '.', filename)
    os.makedirs(app.instance_path or '.', exist_ok=True)
    wb.save(filepath)
    return redirect(f'/download/{filename}')

@app.route('/download/<filename>')
@login_required
def download(filename):
    from flask import send_from_directory
    return send_from_directory(app.instance_path or '.', filename, as_attachment=True)

# ===== DATA IMPORT =====
@app.route('/import')
@login_required
@admin_required
def import_page():
    return render_template('import.html')

@app.route('/import/upload', methods=['POST'])
@login_required
@admin_required
def import_upload():
    entity = request.form.get('entity')
    raw_json = request.form.get('json_data', '').strip()
    if not raw_json:
        flash('Paste some JSON data first.', 'danger')
        return redirect(url_for('import_page'))
    try:
        data = json.loads(raw_json)
        if not isinstance(data, list):
            data = [data]
    except:
        flash('Invalid JSON. Check the format and try again.', 'danger')
        return redirect(url_for('import_page'))

    count = 0
    errors = []

    if entity == 'raw_materials':
        for item in data:
            try:
                bags = item.get('bags', [])
                if item.get('unit') == 'bag':
                    stock = sum(b.get('qty', 0) for b in bags)
                else:
                    stock = float(item.get('currentStock', item.get('current_stock', 0)))
                if RawMaterial.query.filter_by(name=item['name']).first():
                    continue
                rm = RawMaterial(
                    name=item['name'],
                    unit=item.get('unit', 'kg'),
                    bags=json.dumps(bags),
                    current_stock=stock,
                    unit_cost=float(item.get('unitCost', item.get('unit_cost', 0))),
                    reorder_level=float(item.get('reorderLevel', item.get('reorder_level', 0))),
                    description=item.get('description', '')
                )
                db.session.add(rm)
                count += 1
            except Exception as e:
                errors.append(str(e))

    elif entity == 'products':
        for item in data:
            try:
                bags = item.get('bags', [])
                if item.get('unit') == 'bag':
                    stock = sum(b.get('qty', 0) for b in bags)
                else:
                    stock = float(item.get('currentStock', item.get('current_stock', 0)))
                if Product.query.filter_by(name=item['name']).first():
                    continue
                p = Product(
                    name=item['name'],
                    unit=item.get('unit', 'bag'),
                    bags=json.dumps(bags),
                    current_stock=stock,
                    price=float(item.get('price', 0)),
                    description=item.get('description', '')
                )
                db.session.add(p)
                count += 1
            except Exception as e:
                errors.append(str(e))

    elif entity == 'suppliers':
        for item in data:
            try:
                if Supplier.query.filter_by(name=item['name']).first():
                    continue
                s = Supplier(
                    name=item['name'],
                    contact=item.get('contact', ''),
                    phone=item.get('phone', ''),
                    email=item.get('email', ''),
                    address=item.get('address', '')
                )
                db.session.add(s)
                count += 1
            except Exception as e:
                errors.append(str(e))

    elif entity == 'purchases':
        for item in data:
            try:
                rm = RawMaterial.query.get(item.get('rawMaterialId'))
                if not rm:
                    errors.append(f"Raw material ID {item.get('rawMaterialId')} not found")
                    continue
                qty = float(item.get('quantity', 0))
                price = float(item.get('unitPrice', 0))
                total = qty * price
                paid = min(float(item.get('paidAmount', 0)), total)
                p = Purchase(
                    raw_material_id=rm.id,
                    supplier_id=item.get('supplierId'),
                    supplier_name=item.get('supplierName', ''),
                    quantity=qty,
                    unit_price=price,
                    total_cost=total,
                    paid_amount=paid,
                    balance=total - paid,
                    payment_method=item.get('paymentMethod', 'Cash'),
                    payment_date=item.get('paymentDate', ''),
                    invoice_no=item.get('invoiceNo', ''),
                    date=item.get('date', ''),
                    notes=item.get('notes', '')
                )
                db.session.add(p)
                rm.current_stock = (rm.current_stock or 0) + qty
                count += 1
            except Exception as e:
                errors.append(str(e))

    elif entity == 'production':
        for item in data:
            try:
                prod = Product.query.get(item.get('productId'))
                if not prod:
                    errors.append(f"Product ID {item.get('productId')} not found")
                    continue
                qty = float(item.get('quantityProduced', item.get('quantity_produced', 0)))
                ingredients = item.get('ingredients', [])
                for ing in ingredients:
                    mat = RawMaterial.query.get(ing.get('rawMaterialId'))
                    if mat and ing.get('qty', 0) > mat.current_stock:
                        errors.append(f"Insufficient stock of {mat.name}")
                        continue
                    if mat:
                        mat.current_stock = max(0, (mat.current_stock or 0) - ing.get('qty', 0))
                prod.current_stock = (prod.current_stock or 0) + qty
                pp = Production(
                    product_id=prod.id,
                    quantity_produced=qty,
                    ingredients=json.dumps(ingredients),
                    date=item.get('date', today_str())
                )
                db.session.add(pp)
                count += 1
            except Exception as e:
                errors.append(str(e))

    elif entity == 'sales':
        for item in data:
            try:
                prod = Product.query.get(item.get('productId'))
                if not prod:
                    errors.append(f"Product ID {item.get('productId')} not found")
                    continue
                qty = float(item.get('quantity', 0))
                price = float(item.get('unitPrice', 0))
                total = qty * price
                paid = min(float(item.get('paidAmount', 0)), total)
                s = Sale(
                    product_id=prod.id,
                    quantity=qty,
                    unit_price=price,
                    total_amount=total,
                    paid_amount=paid,
                    balance=total - paid,
                    customer_name=item.get('customerName', 'Walk-in'),
                    date=item.get('date', ''),
                    payment_method=item.get('paymentMethod', 'Cash'),
                    payment_date=item.get('paymentDate', ''),
                    notes=item.get('notes', '')
                )
                db.session.add(s)
                prod.current_stock = max(0, (prod.current_stock or 0) - qty)
                count += 1
            except Exception as e:
                errors.append(str(e))

    else:
        flash('Invalid entity type.', 'danger')
        return redirect(url_for('import_page'))

    db.session.commit()
    msg = f'Imported {count} {entity.replace("_", " ")}.'
    if errors:
        msg += f' Errors: {"; ".join(errors[:5])}'
    flash(msg, 'success' if not errors else 'warning')
    return redirect(url_for('import_page'))

# ===== INIT =====
with app.app_context():
    db.create_all()
    # Create/update admin
    admin = User.query.filter_by(username='Sunny').first()
    if not admin:
        admin = User(username='Sunny', role='admin')
        db.session.add(admin)
    admin.set_password('Newphase.sunny11')
    # Create/update viewer
    viewer = User.query.filter_by(username='User').first()
    if not viewer:
        viewer = User(username='User', role='viewer')
        db.session.add(viewer)
    viewer.set_password('Zawadi2026')
    # Remove old viewer if exists
    old_viewer = User.query.filter_by(username='viewer').first()
    if old_viewer:
        db.session.delete(old_viewer)
    # Remove old default admin if exists
    old = User.query.filter_by(username='admin').first()
    if old:
        db.session.delete(old)
    db.session.commit()
    print('Users ready')

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
